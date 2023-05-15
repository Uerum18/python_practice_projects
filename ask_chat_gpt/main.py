import time

import openai
import os
import openpyxl

openai.api_key = 'Insert your openai api key here'
os.chdir(f'{os.getcwd()}/ask_chat_gpt/')
file_path = f'{os.getcwd()}/cat_breeds.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active
row_index = 2


def ask_ai(cat_breed):
    message = {"role": "user",
               "content": f'What is top 3 most common name for cat of {cat_breed} breed?'
                          f'Answer should be in format: name_1, name_2, name_3'}
    chat = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[message]
        )
    reply = chat.choices[0].message.content
    return reply


def get_cat_breeds():
    cat_breeds = []
    for row in sheet.iter_rows(min_row=row_index):
        cat_breeds.append(row[0].value)
    cat_breeds = [breed for breed in cat_breeds if breed]
    return cat_breeds


def write_names_to_file(names):
    for i, row in enumerate(sheet.iter_rows(min_row=row_index)):
        cell = row[1]
        if i < len(names):
            cell.value = names[i]
    workbook.save("cat_names.xlsx")


if __name__ == '__main__':
    cat_names = []
    breeds = get_cat_breeds()
    for breed in breeds:
        answer = ask_ai(breed)
        cat_names.append(answer)
        time.sleep(20)  # Because of openai limitations.
    write_names_to_file(cat_names)
